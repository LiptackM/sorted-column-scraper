"""
unit tests for excel_column_sorted.py functions

notes:
    - requires a test_file & test_corrupt_file (see fixtures) in the same folder
      as this test.  Note if you provide these, also change expected values below
    - test_corrupt-file is jsut a simple text file renamed to a xlsx so its corrupt
    - pytest.ini needs to know the root where these are,
        typically by adding "pythonpath = ."

to run with coverage (coverage suggested, initally this was 87%, when
        only skipping simple I/O in main()):
    > ppytest -rs --cov=excel_column_sorted --cov-report=term-missing 


"""

import pytest
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from excel_column_sorted import process_excel_columns_rw
from excel_column_sorted import remove_before_and_after
from excel_column_sorted import process_and_sort
from excel_column_sorted import print_output
from excel_column_sorted import parse_args


@pytest.fixture
def test_file():
    return "excel_column_sorted_test_content.xlsx"


@pytest.fixture
def test_corrupt_file():
    return "excel_column_sorted_test_corrupt.xlsx"

# ___process_excel_columns_rw ___________________________________________________
# postive test and functional test of process_excel_columns_rw


def test_process_excel_columns_rw_positive(test_file):
    initial_output = []
    for value in process_excel_columns_rw(test_file, 'Test sheet'):
        initial_output.append(value)

    assert initial_output[0] == "Column A"
    assert initial_output[1] == 1
    assert initial_output[2] == 2

# non-existent xls file


def test_process_excel_columns_rw_no_such_file():
    with pytest.raises(FileNotFoundError, match="No such file or directory"):
        list(process_excel_columns_rw(
            'excel_column_sorted_test_nonexistent.xlsx', 'Test sheet'))


# different sheet than the default,
#   this often works because of default so lets make sure a switch is OK
def test_process_excel_columns_rw_different_sheet(test_file):
    # get current sheet
    workbook = None
    filename = test_file
    workbook = load_workbook(filename)
    sheet = workbook.active
    sheet_to_activate = workbook['Second sheet']
    workbook.active = workbook.sheetnames.index(sheet_to_activate.title)
    sheet = workbook.active
    assert sheet.title == 'Second sheet'

    # switch back to desired sheet and read output
    initial_output = []
    for value in process_excel_columns_rw(test_file, 'Test sheet'):
        initial_output.append(value)

    assert initial_output[0] == "Column A"
    assert initial_output[1] == 1
    assert initial_output[2] == 2

# no such sheet exists


def test_process_excel_columns_rw_no_such_sheet(test_file):
    with pytest.raises(KeyError, match="Worksheet Foo does not exist."):
        list(process_excel_columns_rw(test_file, 'Foo'))


# corrupt .xlsx should give a zipfile.BadZipFile: File is not a zip file
def test_process_excel_columns_rw_corrupted_xls(test_corrupt_file):
    with pytest.raises(zipfile.BadZipFile, match="File is not a zip file"):
        # The function call should raise the exception directly for this generator
        list(process_excel_columns_rw(test_corrupt_file, 'Test sheet'))


# ___remove_before_and_after()_____________________________________________

# positive tests, different types of cell contents returned from openpyxl
def test_remove_before_and_after_positive():
    my_list = [
        None,
        "TRUE",
        "FALSE",
        "",
        "text",
        "my Text",
        0.1,
        -10,
        '#Div/0!',
        None,
        0,
        100000000,
        "end",
        None,
    ]

    assert remove_before_and_after(my_list, None, "") == ["TRUE", "FALSE"]
    assert remove_before_and_after(my_list, None, None) == [
        "TRUE", "FALSE", "", "text", "my Text", 0.1, -10, '#Div/0!']
    assert remove_before_and_after(my_list, "my Text", -10) == [0.1]
    assert remove_before_and_after(my_list, "my Text", "foo") == []


# negative tests,different types of cell contents returned from openpyxl
def test_remove_before_and_after_negative():
    my_list = [
        None,
        "TRUE",
        "FALSE",
        "",
        "text",
        "my Text",
        None,
    ]

    assert remove_before_and_after(my_list, "my Text", "foo") == []
    assert remove_before_and_after(my_list, "foo", "0.1") == []
    assert remove_before_and_after(my_list, "foo", "bar") == []


# empty list
def test_remove_before_and_after_empty_list():
    my_list = []

    assert remove_before_and_after(my_list, "my Text", "foo") == []
    assert remove_before_and_after(my_list, "foo", "0.1") == []
    assert remove_before_and_after(my_list, "foo", "bar") == []


# ___process_and_sort()______________________________________________

# positive test of process_and_sort, negative handled in other tests
def test_process_and_sort(test_file):
    output = []
    output = process_and_sort(test_file, 'Column F', 21, 'Test sheet')
    assert output == [1, 2, 3, 4, 5, 6, 7, 8]


# ___print_output()___________________________________________________

# any type in list from excel
def test_print_output_any_type(capsys):
    items = [1, "Hello", 3.14, True]
    print_output(items)
    captured = capsys.readouterr()
    assert "1 (type: int)" in captured.out
    assert "'Hello' (type: str)" in captured.out
    assert "3.14 (type: float)" in captured.out
    assert "True (type: bool)" in captured.out

# empty list


def test_print_output_empty(capsys):
    items = []
    print_output(items)
    captured = capsys.readouterr()
    assert captured.out == ""


# ___parse_args()________________________________________________

# exactly 5 valid arguments
def test_parse_args_correct(test_file):
    argv = ["excel_column_sorted.py", test_file,
            "Test sheet", "Column A", "None"]
    result = parse_args(argv)
    expected = (Path(test_file), "Test sheet", "Column A", "None")
    assert result == expected

# fewer than 5 arguments


def test_parse_args_missing_arguments(test_file, capsys):
    argv = ["excel_column_sorted.py", test_file, "Test sheet"]
    result = parse_args(argv)
    captured = capsys.readouterr()
    assert ("Usage: excel_column_sorted.py" +
            " <full path to excel file> <'sheet name'>" +
            " <'column title'> <'ending cell'," +
            " use None if blank>") in captured.out
    assert result is None

# more than 5 arguments


def test_parse_args_too_many_arguments(test_file):
    argv = ["excel_column_sorted.py", test_file,
            "Test sheet", "Column A", "None", "ExtraArg"]
    result = parse_args(argv)
    expected = (Path(test_file), "Test sheet", "Column A", "None")
    assert result == expected
