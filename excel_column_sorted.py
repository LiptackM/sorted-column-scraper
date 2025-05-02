r"""
Gathers all the entries in a sheet, slices to get one or more consecutive
    columns,and then sorts by value.

Usage: excel_column,py <full path to excel file> <"column title">
    <"ending cell", use None if blank>'.  Most likely the last cell is blank,
    so "ending cell" should be None.  Use quote around Column title if more
    than one word to account for blank

Args:

    <path>  fully qualified path to excel file
    <"column title"> is the title of the column to sort.
        Use quote around column title if more than one word to account for
        blank
    <"ending cell">, use None if blank

Returns: sorted list values, one on each line

PS sample
    python excel_column_sorted.py my.xlsx "Blah" None

Tests
    see \tests\test_excel_column_sorted.py

"""

import sys
from pathlib import Path

from openpyxl import load_workbook


def parse_args(argv):
    """
    takes args and returns new variables in a function, so easily testable
    with pylint
    """
    if len(argv) < 5:
        print("Usage: excel_column_sorted.py" +
              " <full path to excel file> <'sheet name'>" +
              " <'column title'> <'ending cell', use None if blank>")
        return None
    initial_path = argv[1]
    sheet_name = argv[2]
    column_title = argv[3]
    ending_cell = argv[4]
    path = Path(initial_path)  # format properly for all Operating Systems
    return path, sheet_name, column_title, ending_cell


def print_output(items: list[any]) -> None:
    """
    separated out  printing to make main() just initial calls and I/) to make
    testable
    """
    for item in items:
        print(f"{repr(item)} (type: {type(item).__name__})", end='\r\n')


def process_excel_columns_rw(filename, sheet_name):
    """
    Opens an Excel file in read/write mode, iterates through its columns,
    and yields cell values. Ensures the workbook is closed using finally.
    """
    workbook = None
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
        sheet_to_activate = workbook[sheet_name]
        workbook.active = workbook.sheetnames.index(sheet_to_activate.title)
        sheet = workbook.active

        if sheet:
            for col in sheet.iter_cols():
                for cell in col:
                    yield cell.value
    finally:
        if workbook:
            workbook.close()


def remove_before_and_after(
    my_list,
    front_target: str,
    back_target: str
) -> list:
    """
    From a list composed from all cells in an excel sheet, removes
    everything before title of desired column, and then keeps everything
    else up to the ending cell.  Note that back_target is the next instance
    of that value in the list if there are multiples

    Args:
        my_list is all the cells in the sheet by column (eg, [A1, A2, A3, B1,
        B2, B3]).
        Front_target is the title of the column you wish to start at
        back_target is the cell after the last one you wish to keep,
            & is often blank which is noted as None

    Return:  a trimmed list

    Exceptions:
    if non-existence of cell contents that created this list, prints a
    ValueException: "'foo' is not in list" and then "target not found" and
    returns an empty list
    """

    try:
        # make sure if None entered it is a NoneType
        back_target = None if back_target == 'None' else back_target
        front_target = None if front_target == 'None' else front_target

        # entering the same value is undefined so return empty list
        # if front_target == back_target:
        # return []

        front_index = my_list.index(front_target)
        front_cut = my_list[front_index + 1:]

        back_index = front_cut.index(back_target)

        trimmed = front_cut[:back_index]

        return trimmed

    except ValueError as value_error:
        print(value_error)
        print("target not found in sheet")
        return []  # target not found; return empty


def process_and_sort(path, column_title, ending_cell, sheet_name):
    """
    takes args and using these, calls process_excel_columns_rw()
    to create a list, then trims using remove_before_and_after(),
    and finally sorts the list

    Args:
        entries from cmdline

    Return:  final list
    """
    initial_output, processed_list = [], []

    # create new list
    for value in process_excel_columns_rw(path, sheet_name):
        initial_output.append(value)

    # cut the list, then sort
    processed_list = remove_before_and_after(
        initial_output, column_title, ending_cell)
    # debug
    # if processed_list == []:
    #    print("error: processed list is empty")
    processed_list.sort()

    return processed_list


def main():
    """
    basic I/O & function calls only
    """
    final = []

    args: tuple[Path, str, str, str] = parse_args(sys.argv)
    if args is None:
        return None
    path, sheet_name, column_title, ending_cell = args

    final = process_and_sort(path, column_title, ending_cell, sheet_name)

    print_output(final)

    return None


if __name__ == "__main__":
    main()
